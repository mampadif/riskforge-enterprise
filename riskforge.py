import streamlit as st
import google.generativeai as genai
import pandas as pd
import numpy as np
import plotly.graph_objects as go
import plotly.express as px
import stripe
import json
import re
import io
import base64
import hashlib
import os
from datetime import datetime, date, timedelta
from typing import Optional, Dict, Any, List, Tuple
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.lib.units import inch
from rapidfuzz import fuzz, process

# Optional imports – if sentence_transformers fails due to missing torchvision, we disable it
try:
    from sentence_transformers import SentenceTransformer
    from sklearn.metrics.pairwise import cosine_similarity
    EMBEDDING_AVAILABLE = True
except (ImportError, ModuleNotFoundError):
    EMBEDDING_AVAILABLE = False

# =============================================================================
# CONFIGURATION
# =============================================================================
st.set_page_config(page_title="RiskForge Enterprise", page_icon="🛡️", layout="wide")

GEMINI_API_KEY = st.secrets.get("GEMINI_API_KEY")
STRIPE_SECRET_KEY = st.secrets.get("STRIPE_SECRET_KEY")
STRIPE_PRICE_ID_PRO_MONTHLY = st.secrets.get("FORGE_STRIPE_PRICE_ID_PRO_MONTHLY")
STRIPE_PRICE_ID_PRO_ANNUAL = st.secrets.get("FORGE_STRIPE_PRICE_ID_PRO_ANNUAL")
STRIPE_PRICE_ID_ENT_MONTHLY = st.secrets.get("FORGE_STRIPE_PRICE_ID_ENT_MONTHLY")
STRIPE_PRICE_ID_ENT_ANNUAL = st.secrets.get("FORGE_STRIPE_PRICE_ID_ENT_ANNUAL")
APP_URL = st.secrets.get("APP_URL", "https://your-app.streamlit.app")
PRO_UNLOCK_CODE = st.secrets.get("PRO_UNLOCK_CODE", "PRO2025")
ENT_UNLOCK_CODE = st.secrets.get("ENT_UNLOCK_CODE", "ENT2025")

if GEMINI_API_KEY:
    genai.configure(api_key=GEMINI_API_KEY)
    GEMINI_AVAILABLE = True
    ai_model = genai.GenerativeModel("gemini-2.0-flash")
else:
    GEMINI_AVAILABLE = False
    ai_model = None

stripe.api_key = STRIPE_SECRET_KEY if STRIPE_SECRET_KEY else None

# =============================================================================
# SESSION STATE
# =============================================================================
if "tier" not in st.session_state:
    st.session_state.tier = "free"
if "rf_data" not in st.session_state:
    st.session_state.rf_data = None
if "history" not in st.session_state:
    st.session_state.history = []
if "org_name" not in st.session_state:
    st.session_state.org_name = "Your Organisation"
if "report_title" not in st.session_state:
    st.session_state.report_title = "Enterprise Risk Overview"
if "logo_bytes" not in st.session_state:
    st.session_state.logo_bytes = None
if "primary_color" not in st.session_state:
    st.session_state.primary_color = "#0E365C"
if "secondary_color" not in st.session_state:
    st.session_state.secondary_color = "#1A5F7A"
if "board_threshold" not in st.session_state:
    st.session_state.board_threshold = 12
if "default_residual_score" not in st.session_state:
    st.session_state.default_residual_score = 12
if "category_appetite" not in st.session_state:
    st.session_state.category_appetite = {}
if "parser_audit" not in st.session_state:
    st.session_state.parser_audit = None
if "debug_mode" not in st.session_state:
    st.session_state.debug_mode = False
if "force_gemini" not in st.session_state:
    st.session_state.force_gemini = True

def handle_payment_success(plan: str):
    if plan in ("pro_monthly", "pro_annual"):
        st.session_state.tier = "professional"
    elif plan in ("ent_monthly", "ent_annual"):
        st.session_state.tier = "enterprise"

for param in ["success_pro_monthly", "success_pro_annual", "success_ent_monthly", "success_ent_annual"]:
    if param in st.query_params:
        if "pro" in param:
            handle_payment_success("pro_monthly" if "monthly" in param else "pro_annual")
        else:
            handle_payment_success("ent_monthly" if "monthly" in param else "ent_annual")
        st.query_params.clear()

# =============================================================================
# CACHING HELPERS
# =============================================================================
@st.cache_resource
def get_embedding_model():
    if EMBEDDING_AVAILABLE:
        return SentenceTransformer('all-MiniLM-L6-v2')
    return None

def make_json_serializable(obj: Any) -> Any:
    if isinstance(obj, (pd.Timestamp, np.datetime64)):
        return obj.isoformat() if hasattr(obj, 'isoformat') else str(obj)
    if isinstance(obj, (np.integer, np.int64, np.int32)):
        return int(obj)
    if isinstance(obj, (np.floating, np.float64, np.float32)):
        return float(obj)
    if isinstance(obj, np.ndarray):
        return obj.tolist()
    if isinstance(obj, dict):
        return {k: make_json_serializable(v) for k, v in obj.items()}
    if isinstance(obj, (list, tuple)):
        return [make_json_serializable(i) for i in obj]
    return obj

@st.cache_data(ttl=3600)
def cached_parse_file(file_bytes: bytes, file_name: str, default_residual: int) -> Tuple[pd.DataFrame, Dict]:
    return parse_uploaded_file_bytes(file_bytes, file_name, default_residual)

# =============================================================================
# UNIVERSAL PARSER HELPERS
# =============================================================================
def normalize_text(val: Any) -> str:
    if val is None:
        return ""
    try:
        if pd.isna(val):
            return ""
    except Exception:
        pass
    text = str(val).replace("\xa0", " ").strip()
    text = re.sub(r"\s+", " ", text)
    return text

def merged_cell_value(ws, row_idx: int, col_idx: int) -> Any:
    cell = ws.cell(row=row_idx, column=col_idx)
    if cell.value is not None:
        return cell.value

    for merged_range in ws.merged_cells.ranges:
        if (merged_range.min_row <= row_idx <= merged_range.max_row
                and merged_range.min_col <= col_idx <= merged_range.max_col):
            return ws.cell(merged_range.min_row, merged_range.min_col).value
    return None

def parse_due_date(val: Any) -> Optional[date]:
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val

    text = normalize_text(val)
    if not text:
        return None

    for fmt in [
        "%B %d, %Y", "%b %d, %Y", "%b-%y", "%B-%y",
        "%m/%d/%Y", "%m/%d/%y", "%Y-%m-%d", "%d/%m/%Y",
        "%d-%b-%y", "%b-%Y",
    ]:
        try:
            return datetime.strptime(text, fmt).date()
        except Exception:
            pass

    parsed = pd.to_datetime(text, errors="coerce")
    if pd.notna(parsed):
        return parsed.date()
    return None

def clean_division_name(filename: str) -> str:
    name = re.sub(r"\.xlsx$|\.xls$|\.csv$", "", filename, flags=re.IGNORECASE)
    name = re.sub(r"^copy of\s+", "", name, flags=re.IGNORECASE)
    name = name.replace("_", " ").strip()
    name = re.sub(r"\s+", " ", name)
    return name.title() if name else "Unknown Division"

def clean_division_value(val: Any) -> str:
    text = normalize_text(val)
    if not text:
        return ""

    text = re.sub(
        r"^(division\/dept|division|department|dept|directorate|function|unit)\s*[:\-]?\s*",
        "",
        text,
        flags=re.I
    )
    text = re.sub(r"\s+", " ", text).strip(" :-")

    abbrev_map = {
        "hr": "Human Resources",
        "it": "Information Technology",
        "ict": "Information Technology",
        "ist": "Information Systems Technology",
    }

    key = text.lower()
    if key in abbrev_map:
        return abbrev_map[key]

    return text.title()

def looks_like_category(text: str) -> bool:
    categories = {
        "strategic", "financial", "operational", "compliance/legal", "people/hr",
        "ict/cyber", "reputational", "health/safety", "environmental", "legal", "hr", "uncategorised",
    }
    return normalize_text(text).lower() in categories

def detect_explicit_division(ws, header_row: Optional[int] = None, scan_cols: int = 12) -> Optional[str]:
    label_regex = re.compile(
        r"^(division\/dept|division|department|dept|directorate|function|unit)\s*[:\-]?$",
        re.I
    )

    max_row_to_scan = max(1, (header_row - 1) if header_row else 12)
    max_row_to_scan = min(max_row_to_scan, 15)

    for r in range(1, max_row_to_scan + 1):
        for c in range(1, min(scan_cols, ws.max_column) + 1):
            label = normalize_text(merged_cell_value(ws, r, c))
            if not label:
                continue

            if label_regex.match(label):
                candidates = []
                for offset in range(1, 6):
                    candidates.append(merged_cell_value(ws, r, c + offset))
                for offset in range(0, 6):
                    candidates.append(merged_cell_value(ws, r + 1, c + offset))

                for cand in candidates:
                    cand_text = clean_division_value(cand)
                    if not cand_text or len(cand_text) < 3:
                        continue
                    if looks_like_category(cand_text):
                        continue
                    reject_terms = [
                        "risk register", "risk description", "risk definition", "risk category",
                        "impact", "likelihood", "inherent risk", "controls", "owner", "cause",
                        "objective", "updated",
                    ]
                    if any(term in cand_text.lower() for term in reject_terms):
                        continue
                    return cand_text
    return None

def is_valid_division_name(name: str) -> bool:
    if len(name) < 3:
        return False
    if re.match(r"^\d{4}[-/]\d{2}$", name):
        return False
    if re.match(r"^\d+$", name):
        return False
    return True

def extract_division_from_sheet_context(file_name: str, sheet_name: str, header_preview: List[str]) -> str:
    name = sheet_name if sheet_name else file_name
    name = re.sub(r"(?i)\brisk\s*register\b", "", name)
    name = re.sub(r"(?i)\bconsolidated\b", "", name)
    name = re.sub(r"[_\-\s]+", " ", name).strip()
    if name and is_valid_division_name(name):
        return name.title()
    base = os.path.splitext(file_name)[0]
    base = re.sub(r"[_\-\s]+", " ", base).strip()
    if is_valid_division_name(base):
        return base.title()
    return clean_division_name(file_name)

def parse_risk_score(val: Any) -> Optional[float]:
    if val is None:
        return None
    try:
        if pd.isna(val):
            return None
    except Exception:
        pass

    s = normalize_text(val).lower()
    if s in {"#n/a", "#value!", "#ref!", "#name?", "#num!", "#null!", "none", "nan", ""}:
        return None

    impact_map = {"critical": 5, "major": 4, "moderate": 3, "significant": 2, "minor": 1}
    likelihood_map = {"almost certain": 5, "likely": 4, "moderate": 3, "unlikely": 2, "rare": 1}
    inherent_map = {"low": 5, "medium": 10, "high": 16, "critical": 20}

    if s in impact_map:
        return float(impact_map[s])
    if s in likelihood_map:
        return float(likelihood_map[s])
    if s in inherent_map:
        return float(inherent_map[s])

    match = re.search(r"(\d+(?:\.\d+)?)", s)
    if match:
        num = float(match.group(1))
        if 1 <= num <= 25:
            return num
    return None

def parse_control_effectiveness(val: Any) -> Optional[int]:
    if val is None:
        return None
    try:
        if pd.isna(val):
            return None
    except Exception:
        pass

    s = normalize_text(val).lower()
    mapping = {
        "very good": 1, "good": 2, "satisfactory": 3,
        "moderate": 3, "weak": 4, "unsatisfactory": 5, "ineffective": 5,
    }
    for key, num in mapping.items():
        if key in s:
            return num

    match = re.search(r"(\d+)", s)
    if match:
        num = int(match.group(1))
        if 1 <= num <= 5:
            return num
    return None

def infer_category_from_text(title: str, statement: str, cause: str, raw_category: str = "") -> str:
    combined = " ".join([normalize_text(raw_category), normalize_text(title),
                         normalize_text(statement), normalize_text(cause)]).lower()

    if raw_category:
        raw = normalize_text(raw_category)
        if raw:
            return raw

    if any(term in combined for term in ["compliance", "regulation", "labour law", "employment act", "legal"]):
        return "Compliance/Legal"
    if any(term in combined for term in ["staff", "talent", "retention", "morale", "employee", "recruit", "succession", "turnover", "hr"]):
        return "People/HR"
    if any(term in combined for term in ["budget", "funding", "financial", "revenue", "cash"]):
        return "Financial"
    if any(term in combined for term in ["system", "cyber", "data", "technology", "it", "ict", "information technology"]):
        return "ICT/Cyber"
    if any(term in combined for term in ["safety", "injury", "accident", "health"]):
        return "Health/Safety"
    if any(term in combined for term in ["strategy", "strategic"]):
        return "Strategic"
    if any(term in combined for term in ["reputation", "brand", "image"]):
        return "Reputational"
    if any(term in combined for term in ["environment", "climate", "pollution"]):
        return "Environmental"
    return "Operational"

# =============================================================================
# SHEET / HEADER DISCOVERY
# =============================================================================
FIELD_PATTERNS: Dict[str, List[str]] = {
    "risk_no": [r"risk\s*no", r"risk\s*id", r"identifier"],
    "objective": [r"link\s*to\s*objective", r"objective"],
    "risk_name": [r"risk\s*description", r"risk\s*title", r"risk\s*name", r"name\s*of\s*risk", r"^risk$", r"key\s*risk"],
    "risk_statement": [r"risk\s*definition", r"statement", r"risk\s*event", r"risk\s*detail", r"risk\s*narrative", r"description\s*of\s*risk"],
    "cause": [r"cause", r"root\s*cause"],
    "category": [r"risk\s*category", r"category", r"type\s*of\s*risk"],
    "impact_text": [r"impact", r"severity"],
    "likelihood_text": [r"likelihood", r"probability"],
    "inherent_text": [r"inherent\s*risk", r"inherent"],
    "residual_text": [r"residual\s*risk", r"residual"],
    "controls": [r"controls", r"control\s*measure", r"existing\s*controls", r"current\s*controls"],
    "control_effectiveness_text": [r"control\s*effectiveness", r"effectiveness"],
    "owner": [r"risk\s*owner", r"owner", r"accountable", r"action\s*owner"],
    "strategy": [r"risk\s*strategy", r"strategy", r"response"],
    "treatment": [r"risk\s*treatment", r"action\s*plan", r"mitigation", r"treatment", r"management\s*action"],
    "status": [r"status", r"progress", r"quarter\s*status", r"update\s*status"],
    "due_date": [r"treatment\s*due\s*date", r"due\s*date", r"target\s*date", r"completion\s*date"],
}

def is_helper_sheet(sheet_name: str) -> bool:
    s = sheet_name.lower()
    helper_keywords = [
        "boundary", "impact", "likelihood", "effectiveness", "matrix", "lookup",
        "legend", "instruction", "guide", "dashboard", "heatmap", "chart", "pivot",
    ]
    return any(word in s for word in helper_keywords)

def detect_header_row_and_columns(ws, scan_rows: int = 30, max_cols: int = 50) -> Tuple[Optional[int], Dict[str, int], int]:
    best_row = None
    best_map: Dict[str, int] = {}
    best_score = 0

    for row_idx in range(1, min(scan_rows, ws.max_row) + 1):
        current_map: Dict[str, int] = {}
        score = 0

        for col_idx in range(1, min(max_cols, ws.max_column) + 1):
            val = normalize_text(merged_cell_value(ws, row_idx, col_idx)).lower()
            if not val:
                continue

            for field, patterns in FIELD_PATTERNS.items():
                if field in current_map:
                    continue
                if any(re.search(pattern, val) for pattern in patterns):
                    current_map[field] = col_idx
                    score += 1
                    break

        if score > best_score and ("risk_name" in current_map or "risk_statement" in current_map):
            best_row = row_idx
            best_map = current_map
            best_score = score

    return best_row, best_map, best_score

def rank_candidate_sheets(wb) -> List[Dict[str, Any]]:
    candidates: List[Dict[str, Any]] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        header_row, col_map, score = detect_header_row_and_columns(ws)

        name_bonus = 0
        lower_name = sheet_name.lower()
        if "risk" in lower_name:
            name_bonus += 3
        if "register" in lower_name:
            name_bonus += 3
        if "monitor" in lower_name:
            name_bonus += 2
        if "quarter" in lower_name:
            name_bonus += 2
        if is_helper_sheet(sheet_name):
            name_bonus -= 4

        total_score = score + name_bonus

        candidates.append({
            "sheet_name": sheet_name,
            "header_row": header_row,
            "column_map": col_map,
            "header_score": score,
            "total_score": total_score,
        })

    candidates.sort(key=lambda x: x["total_score"], reverse=True)
    return candidates

# =============================================================================
# ROW ACCEPTANCE SCORING (INTELLIGENCE LAYER)
# =============================================================================
def is_valid_risk_record(row: Dict[str, Any]) -> Tuple[bool, int, str]:
    """Returns (is_valid, confidence_score, reason)."""
    evidence = 0
    reasons = []

    if row.get("risk_name") and len(row["risk_name"]) > 3:
        evidence += 2
        reasons.append("has_risk_name")
    else:
        reasons.append("missing_risk_name")

    if row.get("risk_statement") and len(row["risk_statement"]) > 20:
        evidence += 2
        reasons.append("has_risk_statement")
    else:
        reasons.append("missing_or_short_statement")

    if row.get("cause"):
        evidence += 1
        reasons.append("has_cause")
    if row.get("controls"):
        evidence += 1
        reasons.append("has_controls")
    if row.get("owner") and row["owner"].lower() != "not assigned":
        evidence += 1
        reasons.append("has_owner")
    if row.get("category"):
        evidence += 1
        reasons.append("has_category")
    if row.get("impact_score") is not None:
        evidence += 1
        reasons.append("has_impact")
    if row.get("likelihood_score") is not None:
        evidence += 1
        reasons.append("has_likelihood")
    if row.get("treatment_plan"):
        evidence += 1
        reasons.append("has_treatment")

    is_valid = evidence >= 5
    reason_str = " | ".join(reasons)
    return is_valid, evidence, reason_str

def merge_continuation_rows(rows: List[Dict]) -> List[Dict]:
    """Merge rows that are continuations (no risk_name/statement) into previous risk."""
    merged = []
    current = None
    for row in rows:
        if row.get("risk_name") or row.get("risk_statement"):
            if current:
                merged.append(current)
            current = row.copy()
        else:
            if current:
                current["risk_statement"] += " " + (row.get("risk_statement") or "")
                current["cause"] += " " + (row.get("cause") or "")
                current["controls"] += " " + (row.get("controls") or "")
                current["treatment_plan"] += " " + (row.get("treatment_plan") or "")
    if current:
        merged.append(current)
    return merged

def compute_scores(
    raw_impact: Any,
    raw_likelihood: Any,
    raw_inherent: Any,
    raw_residual: Any,
    raw_control_effectiveness: Any,
    default_residual: int
) -> Tuple[int, int, Optional[float], Optional[float], Optional[int], str]:
    impact_score = parse_risk_score(raw_impact)
    likelihood_score = parse_risk_score(raw_likelihood)
    inherent_score = parse_risk_score(raw_inherent)
    residual_score = parse_risk_score(raw_residual)

    control_eff_numeric = parse_control_effectiveness(raw_control_effectiveness)
    control_eff_text = normalize_text(raw_control_effectiveness) or "Not rated"

    # Inherent
    if inherent_score is not None and inherent_score > 5:
        inherent = float(inherent_score)
    elif impact_score is not None and likelihood_score is not None:
        inherent = float(impact_score * likelihood_score)
    elif impact_score is not None:
        inherent = float(impact_score * 5)
    elif likelihood_score is not None:
        inherent = float(likelihood_score * 5)
    else:
        inherent = float(default_residual)

    inherent = min(25, max(1, round(inherent)))

    # Residual
    if residual_score is not None and residual_score > 5:
        residual = float(residual_score)
    elif control_eff_numeric is not None:
        residual = round(inherent * (control_eff_numeric / 5.0))
    else:
        residual = inherent

    residual = min(25, max(1, round(residual)))

    return residual, inherent, impact_score, likelihood_score, control_eff_numeric, control_eff_text

def parse_structured_sheet(
    ws,
    sheet_name: str,
    col_map: Dict[str, int],
    header_row: int,
    file_name: str,
    default_residual: int
) -> Tuple[pd.DataFrame, Dict[str, Any]]:
    raw_rows: List[Dict[str, Any]] = []
    row_audit: List[Dict[str, Any]] = []

    header_preview = [normalize_text(merged_cell_value(ws, header_row, c)) for c in range(1, min(ws.max_column, 20) + 1)]

    explicit_division = detect_explicit_division(ws, header_row=header_row)
    if explicit_division:
        division_name = explicit_division
    else:
        division_name = extract_division_from_sheet_context(file_name, sheet_name, header_preview)

    def get_field(row_idx: int, field_name: str) -> Any:
        col_idx = col_map.get(field_name)
        if not col_idx:
            return None
        return merged_cell_value(ws, row_idx, col_idx)

    blank_streak = 0
    start_row = header_row + 1

    for row_idx in range(start_row, min(ws.max_row + 1, start_row + 300)):
        raw_risk_no = get_field(row_idx, "risk_no")
        raw_objective = get_field(row_idx, "objective")
        raw_risk_name = get_field(row_idx, "risk_name")
        raw_risk_statement = get_field(row_idx, "risk_statement")
        raw_cause = get_field(row_idx, "cause")
        raw_category = get_field(row_idx, "category")
        raw_impact_text = get_field(row_idx, "impact_text")
        raw_likelihood_text = get_field(row_idx, "likelihood_text")
        raw_inherent_text = get_field(row_idx, "inherent_text")
        raw_residual_text = get_field(row_idx, "residual_text")
        raw_controls = get_field(row_idx, "controls")
        raw_control_eff = get_field(row_idx, "control_effectiveness_text")
        raw_owner = get_field(row_idx, "owner")
        raw_strategy = get_field(row_idx, "strategy")
        raw_treatment = get_field(row_idx, "treatment")
        raw_status = get_field(row_idx, "status")
        raw_due_date = get_field(row_idx, "due_date")

        risk_no = normalize_text(raw_risk_no)
        objective = normalize_text(raw_objective)
        risk_name = normalize_text(raw_risk_name)
        risk_statement = normalize_text(raw_risk_statement)
        cause = normalize_text(raw_cause)
        category_raw = normalize_text(raw_category)
        controls = normalize_text(raw_controls)
        owner = normalize_text(raw_owner)
        strategy = normalize_text(raw_strategy)
        treatment = normalize_text(raw_treatment)
        status = normalize_text(raw_status) or "Active"
        due_date_raw = normalize_text(raw_due_date)

        # Skip empty rows
        combined_gate = " ".join([risk_no, risk_name, risk_statement, cause, controls]).strip()
        if not combined_gate:
            blank_streak += 1
            if blank_streak >= 12:
                break
            continue
        blank_streak = 0

        # Skip obvious metadata/headers
        lower_gate = combined_gate.lower()
        metadata_indicators = [
            "risk description", "risk definition", "link to objective", "control effectiveness",
            "treatment due date", "risk strategy", "risk owner", "impact (1-5)", "likelihood score",
            "inherent risk", "residual risk", "division/dept", "what is in place to prevent",
        ]
        if any(ind in lower_gate for ind in metadata_indicators):
            continue

        if risk_name.startswith("=") or risk_statement.startswith("="):
            continue

        # Compute scores
        residual, inherent, impact_score, likelihood_score, control_eff_numeric, control_eff_text = compute_scores(
            raw_impact=raw_impact_text,
            raw_likelihood=raw_likelihood_text,
            raw_inherent=raw_inherent_text,
            raw_residual=raw_residual_text,
            raw_control_effectiveness=raw_control_eff,
            default_residual=default_residual,
        )

        parsed_due_date = parse_due_date(raw_due_date)
        category = infer_category_from_text(
            title=risk_name,
            statement=risk_statement,
            cause=cause,
            raw_category=category_raw,
        )

        raw_rows.append({
            "division": division_name,
            "risk_no": risk_no,
            "objective_link": objective,
            "risk_name": risk_name,
            "risk_statement": risk_statement,
            "cause": cause,
            "category": category,
            "residual_score": residual,
            "inherent_score": inherent,
            "owner": owner or "Not assigned",
            "status": status,
            "due_date": parsed_due_date,
            "due_date_raw": due_date_raw,
            "control_effectiveness": control_eff_text,
            "control_effectiveness_numeric": control_eff_numeric,
            "impact_score": impact_score,
            "likelihood_score": likelihood_score,
            "controls": controls,
            "strategy": strategy or "Treat",
            "treatment_plan": treatment,
            "source_file": file_name,
            "source_sheet": sheet_name,
            "source_row": row_idx,
        })

    # Merge continuation rows
    raw_rows = merge_continuation_rows(raw_rows)

    # Validate and score each row
    accepted_risks = []
    for row in raw_rows:
        is_valid, confidence, reason = is_valid_risk_record(row)
        if is_valid:
            row["acceptance_score"] = confidence
            row["acceptance_reason"] = reason
            row["parser_confidence"] = 0.95
            accepted_risks.append(row)
            row_audit.append({
                "row": row["source_row"],
                "risk_name": row["risk_name"],
                "status": "accepted",
                "confidence": confidence,
                "reason": reason,
            })
        else:
            row_audit.append({
                "row": row["source_row"],
                "risk_name": row["risk_name"],
                "status": "rejected",
                "confidence": confidence,
                "reason": reason,
            })

    debug = {
        "sheet_used": sheet_name,
        "header_row": header_row,
        "column_map": col_map,
        "rows_scanned": len(raw_rows) + len(row_audit) - len(accepted_risks),  # approximate
        "rows_accepted": len(accepted_risks),
        "acceptance_rate": round(len(accepted_risks) / max(1, len(raw_rows)) * 100, 1),
        "row_audit_preview": row_audit[:100],
    }

    return pd.DataFrame(accepted_risks), debug

# =============================================================================
# UNIVERSAL PARSER ENTRY POINT
# =============================================================================
def parse_structured_risk_register(
    file_bytes: bytes,
    file_name: str,
    default_residual: int
) -> Tuple[pd.DataFrame, Dict[str, Any]]:
    debug_info: Dict[str, Any] = {
        "parser": "intelligent_structured_v3",
        "candidate_sheets": [],
        "selected_sheets": [],
        "error": None,
    }

    try:
        wb = load_workbook(io.BytesIO(file_bytes), data_only=True)

        candidates = rank_candidate_sheets(wb)
        debug_info["candidate_sheets"] = candidates[:10]

        strong_candidates = [
            c for c in candidates
            if c["header_row"] is not None
            and c["header_score"] >= 2
            and len(c["column_map"]) >= 4
            and ("risk_name" in c["column_map"] or "risk_statement" in c["column_map"])
        ]

        if not strong_candidates:
            debug_info["error"] = "No structured risk register sheet detected"
            return pd.DataFrame(), debug_info

        # Parse only the best sheet (avoid over-extraction)
        best = strong_candidates[0]
        debug_info["selected_sheets"] = [best["sheet_name"]]

        ws = wb[best["sheet_name"]]
        df_sheet, parse_debug = parse_structured_sheet(
            ws=ws,
            sheet_name=best["sheet_name"],
            col_map=best["column_map"],
            header_row=best["header_row"],
            file_name=file_name,
            default_residual=default_residual,
        )
        debug_info.update(parse_debug)

        if not df_sheet.empty:
            return df_sheet, debug_info

        debug_info["error"] = "No valid risks parsed after acceptance scoring"
        return pd.DataFrame(), debug_info

    except Exception as exc:
        debug_info["error"] = str(exc)
        return pd.DataFrame(), debug_info

# =============================================================================
# SIMPLE FALLBACK PARSER
# =============================================================================
def simple_fallback_parser(
    file_bytes: bytes,
    file_name: str,
    default_residual: int
) -> Tuple[pd.DataFrame, Dict[str, Any]]:
    risks: List[Dict[str, Any]] = []
    debug_info: Dict[str, Any] = {"parser": "simple_fallback", "cells_scanned": 0, "risks_found": 0}

    try:
        wb = load_workbook(io.BytesIO(file_bytes), data_only=True)

        for sheet_name in wb.sheetnames:
            if is_helper_sheet(sheet_name):
                continue

            ws = wb[sheet_name]

            for row_idx in range(1, min(ws.max_row + 1, 200)):
                for col_idx in range(1, min(ws.max_column + 1, 30)):
                    val = merged_cell_value(ws, row_idx, col_idx)
                    debug_info["cells_scanned"] += 1

                    if isinstance(val, str):
                        text = normalize_text(val)
                        if len(text) > 40 and not text.startswith("="):
                            metadata_indicators = [
                                "risk monitoring", "period of assessment", "department/ unit",
                                "description of control", "overall effectiveness", "control categorisation",
                                "action plan", "residual risk", "risk strategy", "inherent risk",
                                "risk owner", "impact rating", "likelihood rating", "root cause",
                                "consequences", "start date", "target completion", "processed by",
                                "approved by", "recommended by", "signature", "link to objective",
                                "risk no", "division/ dept",
                            ]
                            if not any(ind in text.lower() for ind in metadata_indicators):
                                row = {
                                    "division": clean_division_name(file_name),
                                    "risk_no": "",
                                    "objective_link": "",
                                    "risk_name": text[:80],
                                    "risk_statement": text[:500],
                                    "cause": "",
                                    "category": "Uncategorised",
                                    "residual_score": default_residual,
                                    "inherent_score": min(25, default_residual + 3),
                                    "owner": "Not assigned",
                                    "status": "Active",
                                    "due_date": None,
                                    "due_date_raw": "",
                                    "control_effectiveness": "Not rated",
                                    "control_effectiveness_numeric": None,
                                    "impact_score": None,
                                    "likelihood_score": None,
                                    "controls": "",
                                    "strategy": "Treat",
                                    "treatment_plan": "",
                                    "source_file": file_name,
                                    "source_sheet": sheet_name,
                                    "source_row": row_idx,
                                }
                                is_valid, conf, reason = is_valid_risk_record(row)
                                if is_valid:
                                    row["acceptance_score"] = conf
                                    row["acceptance_reason"] = reason
                                    row["parser_confidence"] = 0.60
                                    risks.append(row)
                                    debug_info["risks_found"] += 1
                                    if len(risks) >= 20:
                                        break
                if len(risks) >= 20:
                    break
            if len(risks) >= 20:
                break

    except Exception as exc:
        debug_info["error"] = str(exc)

    if risks:
        return pd.DataFrame(risks), debug_info
    return pd.DataFrame(), debug_info

# =============================================================================
# GEMINI FALLBACK
# =============================================================================
def gemini_extract_risks(file_bytes: bytes, file_name: str, default_residual: int) -> Tuple[pd.DataFrame, Dict]:
    if not GEMINI_AVAILABLE:
        return pd.DataFrame(), {"error": "Gemini not available"}
    try:
        xls = pd.ExcelFile(io.BytesIO(file_bytes))
        all_content = []
        for sheet in xls.sheet_names:
            if sheet.lower() in ["boundaries", "impact", "likelihood", "effectiveness", "risk matrix"]:
                continue
            df = pd.read_excel(io.BytesIO(file_bytes), sheet_name=sheet, header=None)
            sheet_str = f"\n[SHEET: {sheet}]\n"
            for i in range(min(50, df.shape[0])):
                row_vals = []
                for j in range(min(20, df.shape[1])):
                    val = df.iat[i, j]
                    if pd.notna(val) and str(val).strip():
                        row_vals.append(str(val).strip())
                if row_vals:
                    sheet_str += f"Row {i+1}: {' | '.join(row_vals)}\n"
            all_content.append(sheet_str)
        content = "\n".join(all_content)[:8000]
        prompt = f"""
You are a risk extraction engine. Extract ALL risks as a JSON list.
Each risk must have: risk_statement, risk_name, owner, residual_score (1-25).
Return ONLY valid JSON.
Content:
{content}
"""
        response = ai_model.generate_content(prompt)
        text = response.text.strip()
        if text.startswith("```json"): text = text[7:]
        if text.startswith("```"): text = text[3:]
        if text.endswith("```"): text = text[:-3]
        data = json.loads(text)
        risks = data.get("risks", []) if isinstance(data, dict) else (data if isinstance(data, list) else [])
        rows = []
        for r in risks:
            stmt = r.get("risk_statement", "")
            if not stmt or len(stmt) < 15:
                continue
            row = {
                "division": clean_division_name(file_name),
                "risk_name": r.get("risk_name", stmt[:50]),
                "risk_statement": stmt[:500],
                "category": "Uncategorised",
                "residual_score": min(25, max(1, int(r.get("residual_score", default_residual)))),
                "inherent_score": min(25, int(r.get("residual_score", default_residual)) + 3),
                "owner": r.get("owner", "Not assigned"),
                "status": "Active",
                "due_date": None,
                "control_effectiveness": "Not rated",
                "impact_score": r.get("impact_score"),
                "likelihood_score": r.get("likelihood_score"),
                "source_file": file_name,
                "source_sheet": "gemini",
                "source_row": 0,
                "acceptance_score": 5,
                "acceptance_reason": "gemini_extracted",
                "parser_confidence": 0.80,
            }
            rows.append(row)
        if rows:
            return pd.DataFrame(rows), {"extracted": len(rows), "method": "gemini"}
        return pd.DataFrame(), {"error": "No risks found"}
    except Exception as e:
        return pd.DataFrame(), {"error": f"Gemini failed: {str(e)}"}

# =============================================================================
# FINAL DISPATCHER
# =============================================================================
def parse_uploaded_file_bytes(
    file_bytes: bytes,
    file_name: str,
    default_residual: int
) -> Tuple[pd.DataFrame, Dict[str, Any]]:
    # 1. Intelligent structured parser
    df, debug = parse_structured_risk_register(file_bytes, file_name, default_residual)
    if not df.empty:
        st.success(f"✅ Extracted {len(df)} risks (acceptance rate {debug.get('acceptance_rate', 0)}%)")
        return df, debug

    # 2. Simple fallback
    df, fallback_debug = simple_fallback_parser(file_bytes, file_name, default_residual)
    if not df.empty:
        st.warning(f"⚠️ Simple fallback extracted {len(df)} potential risks")
        return df, fallback_debug

    # 3. Gemini fallback
    if GEMINI_AVAILABLE and st.session_state.force_gemini:
        df, gemini_debug = gemini_extract_risks(file_bytes, file_name, default_residual)
        if not df.empty:
            st.success(f"✅ Gemini extracted {len(df)} risks")
            return df, gemini_debug
        return pd.DataFrame(), gemini_debug

    return pd.DataFrame(), debug

# =============================================================================
# ENTERPRISE CLUSTER CONSOLIDATION
# =============================================================================
def build_enterprise_clusters(df: pd.DataFrame) -> pd.DataFrame:
    """Group similar risks into enterprise clusters."""
    if df.empty:
        return df

    df = df.copy()
    df["cluster_id"] = None
    df["cluster_label"] = None
    df["cluster_primary_statement"] = None

    statements = df["risk_statement"].fillna("").tolist()
    names = df["risk_name"].fillna("").tolist()

    clusters = []
    used = set()

    for i, stmt in enumerate(statements):
        if i in used:
            continue
        # Find similar statements
        matches = process.extract(stmt, statements, scorer=fuzz.token_sort_ratio, limit=10)
        cluster_indices = [i]
        for match in matches:
            if match[1] >= 75 and match[2] not in used:
                cluster_indices.append(match[2])
                used.add(match[2])
        used.update(cluster_indices)
        clusters.append(cluster_indices)

    for cluster_id, indices in enumerate(clusters):
        cluster_label = names[indices[0]]
        primary_stmt = statements[indices[0]]
        for idx in indices:
            df.at[idx, "cluster_id"] = cluster_id
            df.at[idx, "cluster_label"] = cluster_label
            df.at[idx, "cluster_primary_statement"] = primary_stmt

    return df

def parse_all_files(uploaded_files, tier: str, default_residual: int) -> Tuple[pd.DataFrame, List[Dict]]:
    all_risks = []
    all_debug = []
    total_scanned = 0
    total_accepted = 0

    for file in uploaded_files:
        df, debug = cached_parse_file(file.getvalue(), file.name, default_residual)
        all_debug.append(debug)
        total_scanned += debug.get("rows_scanned", 0)
        total_accepted += len(df)
        if not df.empty:
            all_risks.append(df)

    if not all_risks:
        return pd.DataFrame(), all_debug

    df_all = pd.concat(all_risks, ignore_index=True)

    # Apply enterprise clustering
    df_all = build_enterprise_clusters(df_all)

    # Deduplicate exact matches
    before_dedupe = len(df_all)
    df_all["_stmt_norm"] = df_all["risk_statement"].fillna("").apply(lambda x: re.sub(r"[^\w\s]", "", x.lower()).strip())
    df_all = df_all.drop_duplicates(subset=["_stmt_norm"], keep="first")
    df_all = df_all.drop(columns=["_stmt_norm"])

    df_all["residual_level"] = df_all["residual_score"].apply(
        lambda x: "Critical" if x >= 20 else "High" if x >= 12 else "Medium" if x >= 6 else "Low"
    )

    if tier == "free":
        df_all = df_all.head(10)

    st.session_state.parser_audit = {
        "total_files": len(uploaded_files),
        "total_risks": len(df_all),
        "rows_scanned": total_scanned,
        "rows_accepted": total_accepted,
        "acceptance_rate": round(total_accepted / max(1, total_scanned) * 100, 1),
        "duplicates_removed": before_dedupe - len(df_all),
        "clusters_formed": df_all["cluster_id"].nunique() if "cluster_id" in df_all.columns else 0,
    }

    return df_all, all_debug

# =============================================================================
# AI FUNCTIONS
# =============================================================================
def ai_executive_briefing(snapshot: Dict, correlations: List[Dict], recommendations: Dict, company: str) -> str:
    if not GEMINI_AVAILABLE:
        return ""

    context = f"""
    Company: {company}
    Health Score: {snapshot.get('enterprise_health_score', 0)}/100
    Total Risks: {snapshot.get('total_risks', 0)}
    Critical+High: {snapshot.get('critical_count', 0) + snapshot.get('high_count', 0)}
    Top Division: {snapshot.get('top_division', 'N/A')} ({snapshot.get('top_division_pct', 0)}% of exposure)
    Appetite Breached: {snapshot.get('pct_breached', 0)}%
    Treatment Confidence: {snapshot.get('treatment_confidence', 0)}%

    Cross-Division Correlations Found: {len(correlations)}
    """

    if correlations:
        context += "\nTop Correlated Risks Across Divisions:\n"
        for corr in correlations[:3]:
            context += f"- {corr['risk_a']['name']} ({corr['risk_a']['division']}) ↔ {corr['risk_b']['name']} ({corr['risk_b']['division']}) - {corr['similarity']}% similar\n"

    if recommendations:
        context += f"\nThreshold Recommendation: {recommendations.get('global', {}).get('recommended', 'N/A')} (current: {snapshot.get('threshold', 12)})\n"

    prompt = f"""
    You are a Chief Risk Officer writing a board‑level executive briefing.
    Based on the data below, write a concise but insightful 4‑5 sentence briefing that:
    - Highlights the most critical enterprise health indicators.
    - Identifies cross‑divisional risk patterns and what they mean.
    - Suggests 1‑2 actionable focus areas for the next quarter.
    - Uses professional, confident language suitable for the board.

    Data:
    {context}

    Briefing:
    """

    try:
        response = ai_model.generate_content(prompt)
        return response.text.strip()
    except:
        return ""

# =============================================================================
# INTELLIGENCE ENGINE
# =============================================================================
def calculate_weighted_treatment_confidence(row: pd.Series) -> float:
    score = 0.0
    weights = {"owner": 0.3, "due_date": 0.25, "status": 0.25, "control": 0.2}
    if row.get("owner") and row["owner"].lower() != "not assigned":
        score += weights["owner"] * 100
    due = row.get("due_date")
    if due and isinstance(due, (date, pd.Timestamp)):
        if due >= date.today():
            score += weights["due_date"] * 100
        else:
            score += weights["due_date"] * 30
    elif due:
        score += weights["due_date"] * 50
    else:
        score += weights["due_date"] * 20
    status = str(row.get("status", "")).lower()
    if "completed" in status or "closed" in status:
        score += weights["status"] * 100
    elif "on track" in status or "active" in status or "in progress" in status:
        score += weights["status"] * 70
    elif "overdue" in status or "delayed" in status:
        score += weights["status"] * 30
    else:
        score += weights["status"] * 50
    eff = str(row.get("control_effectiveness", "")).lower()
    if "effective" in eff or "strong" in eff:
        score += weights["control"] * 100
    elif "partial" in eff or "moderate" in eff:
        score += weights["control"] * 60
    elif "weak" in eff or "ineffective" in eff:
        score += weights["control"] * 20
    else:
        score += weights["control"] * 40
    return round(score, 1)

def appetite_band(score: float, threshold: int, category: str = "", category_appetite: Dict = None) -> str:
    if pd.isna(score):
        return "unknown"
    if category_appetite and category in category_appetite:
        threshold = category_appetite[category]
    if score >= threshold + 4:
        return "critical breach"
    if score >= threshold:
        return "breached"
    if score >= threshold - 4:
        return "near appetite"
    return "within appetite"

def calculate_enterprise_health_score(df: pd.DataFrame) -> float:
    if df.empty:
        return 0.0
    score = 100.0
    critical_pct = (df["residual_level"] == "Critical").mean() * 100
    score -= critical_pct * 0.8
    unassigned_pct = (df["owner"].astype(str).str.lower() == "not assigned").mean() * 100
    score -= unassigned_pct * 0.5
    high_pct = (df["residual_level"] == "High").mean() * 100
    score -= high_pct * 0.3
    return max(0.0, round(score, 1))

def detect_emerging_themes(df: pd.DataFrame) -> List[str]:
    if df.empty:
        return []
    theme_keywords = {
        "Cyber & Data Security": ["cyber", "ransomware", "data breach", "hacking", "phishing", "malware"],
        "Supply Chain": ["supplier", "vendor", "third party", "outsource", "dependency"],
        "Regulatory Change": ["regulation", "compliance", "legislation", "law change"],
        "Talent & Workforce": ["staff", "skills", "turnover", "recruitment", "retention", "morale", "attrition"],
        "Technology Disruption": ["digital", "automation", "ai", "artificial intelligence", "legacy system"],
        "Financial Pressure": ["budget", "funding", "capital", "investment", "cash flow", "liquidity"]
    }
    themes = []
    for theme, keywords in theme_keywords.items():
        count = sum(1 for stmt in df["risk_statement"].fillna("").astype(str) if any(kw in stmt.lower() for kw in keywords))
        if count >= 2 and (count / len(df)) >= 0.08:
            themes.append(theme)
    return themes

def find_cross_division_correlations(df: pd.DataFrame, threshold: float = 0.75) -> List[Dict]:
    if df.empty or len(df["division"].unique()) < 2:
        return []

    correlations = []
    statements = df["risk_statement"].fillna("").tolist()
    divisions = df["division"].tolist()
    risk_names = df["risk_name"].tolist()

    for i in range(len(statements)):
        for j in range(i + 1, len(statements)):
            if divisions[i] == divisions[j]:
                continue

            text_i = f"{risk_names[i]} {statements[i]}"
            text_j = f"{risk_names[j]} {statements[j]}"

            similarity = fuzz.ratio(text_i.lower(), text_j.lower()) / 100.0
            if similarity >= threshold:
                correlations.append({
                    "risk_a": {"name": risk_names[i], "division": divisions[i], "statement": statements[i][:200]},
                    "risk_b": {"name": risk_names[j], "division": divisions[j], "statement": statements[j][:200]},
                    "similarity": round(similarity * 100, 1)
                })

    correlations.sort(key=lambda x: x["similarity"], reverse=True)
    return correlations[:10]

def recommend_appetite_thresholds(df: pd.DataFrame, current_threshold: int) -> Dict[str, Any]:
    if df.empty:
        return {}

    recommendations = {}
    overall_q75 = df["residual_score"].quantile(0.75)

    recommendations["global"] = {
        "current": current_threshold,
        "recommended": int(min(25, max(6, round(overall_q75)))),
        "reason": f"75th percentile of residual scores is {overall_q75:.1f}"
    }

    for category in df["category"].unique():
        cat_df = df[df["category"] == category]
        if len(cat_df) >= 3:
            cat_q75 = cat_df["residual_score"].quantile(0.75)
            recommendations[category] = {
                "recommended": int(min(25, max(6, round(cat_q75)))),
                "reason": f"75th percentile for {category} risks is {cat_q75:.1f}"
            }

    return recommendations

def analyze_trends(current: Dict, previous: Dict) -> Dict[str, Any]:
    if not previous:
        return {}

    trends = {}
    health_delta = current.get("enterprise_health_score", 0) - previous.get("enterprise_health_score", 0)
    trends["health_trend"] = "improving" if health_delta > 0 else "declining" if health_delta < 0 else "stable"
    trends["health_delta"] = health_delta

    breach_delta = current.get("pct_breached", 0) - previous.get("pct_breached", 0)
    trends["breach_trend"] = "worsening" if breach_delta > 0 else "improving" if breach_delta < 0 else "stable"

    prev_cat = previous.get("category_exposure", {})
    curr_cat = current.get("category_exposure", {})
    growth = {}
    for cat, exp in curr_cat.items():
        prev_exp = prev_cat.get(cat, 0)
        if prev_exp > 0:
            growth[cat] = (exp - prev_exp) / prev_exp * 100

    if growth:
        fastest = max(growth, key=growth.get)
        trends["fastest_growing_category"] = {"category": fastest, "growth_pct": round(growth[fastest], 1)}

    return trends

def build_intelligence_snapshot(df: pd.DataFrame, threshold: int, category_appetite: Dict = None) -> Dict[str, Any]:
    if df.empty:
        return {}
    snapshot = {}
    snapshot["critical_count"] = int((df["residual_level"] == "Critical").sum())
    snapshot["high_count"] = int((df["residual_level"] == "High").sum())
    snapshot["avg_residual"] = round(df["residual_score"].mean(), 1)
    snapshot["avg_inherent"] = round(df["inherent_score"].mean(), 1)
    exposure_by_div = df.groupby("division")["residual_score"].sum().sort_values(ascending=False)
    if not exposure_by_div.empty:
        snapshot["top_division"] = exposure_by_div.index[0]
        snapshot["top_division_pct"] = round((exposure_by_div.iloc[0] / exposure_by_div.sum()) * 100, 1)
        snapshot["division_exposure"] = exposure_by_div.head(5).to_dict()
    else:
        snapshot["top_division"] = "N/A"
        snapshot["top_division_pct"] = 0
        snapshot["division_exposure"] = {}
    exposure_by_cat = df.groupby("category")["residual_score"].sum().sort_values(ascending=False)
    snapshot["category_exposure"] = exposure_by_cat.head(5).to_dict()
    df["appetite_band"] = df.apply(lambda row: appetite_band(row["residual_score"], threshold, row.get("category", ""), category_appetite), axis=1)
    snapshot["pct_within_appetite"] = round((df["appetite_band"] == "within appetite").mean() * 100, 1)
    snapshot["pct_near_appetite"] = round((df["appetite_band"] == "near appetite").mean() * 100, 1)
    snapshot["pct_breached"] = round((df["appetite_band"].isin(["breached", "critical breach"])).mean() * 100, 1)
    snapshot["emerging_themes"] = detect_emerging_themes(df)
    snapshot["ownership_coverage"] = round((df["owner"].astype(str).str.lower() != "not assigned").mean() * 100, 1)
    snapshot["enterprise_health_score"] = calculate_enterprise_health_score(df)
    df["treatment_confidence"] = df.apply(calculate_weighted_treatment_confidence, axis=1)
    snapshot["treatment_confidence"] = round(df["treatment_confidence"].mean(), 1)
    snapshot["total_risks"] = len(df)
    snapshot["board_risks"] = df.nlargest(5, "residual_score")[["risk_name", "division", "residual_score", "owner", "category"]].to_dict("records")
    snapshot["risks_list"] = df[["risk_name", "residual_score", "owner", "division"]].to_dict("records")
    snapshot["threshold"] = threshold
    return snapshot

def compare_snapshots(current: Dict, previous: Dict) -> Dict[str, Any]:
    if not previous or not current:
        return {}
    current_risks = {r["risk_name"]: r for r in current.get("risks_list", [])}
    previous_risks = {r["risk_name"]: r for r in previous.get("risks_list", [])}
    return {
        "new_risks": [current_risks[n] for n in current_risks if n not in previous_risks],
        "closed_risks": [previous_risks[n] for n in previous_risks if n not in current_risks],
        "worsened_risks": [{"name": n, "delta": current_risks[n]["residual_score"] - previous_risks[n]["residual_score"]}
                           for n in current_risks if n in previous_risks and current_risks[n]["residual_score"] > previous_risks[n]["residual_score"] + 1],
        "improved_risks": [{"name": n, "delta": current_risks[n]["residual_score"] - previous_risks[n]["residual_score"]}
                           for n in current_risks if n in previous_risks and current_risks[n]["residual_score"] < previous_risks[n]["residual_score"] - 1],
        "health_delta": current.get("enterprise_health_score", 0) - previous.get("enterprise_health_score", 0),
        "appetite_delta": current.get("pct_breached", 0) - previous.get("pct_breached", 0),
    }

def generate_board_narrative(snapshot: Dict, comparison: Dict, threshold: int, company: str, report_title: str, ai_summary: str = "") -> str:
    narrative = []
    narrative.append(f"# {report_title}")
    narrative.append(f"**Organization:** {company}")
    narrative.append(f"**Date:** {datetime.now().strftime('%B %d, %Y')}")
    narrative.append(f"**Reporting Period:** Q{((datetime.now().month-1)//3)+1} {datetime.now().year}")
    narrative.append("")
    if ai_summary:
        narrative.append(f"**AI Executive Summary:** {ai_summary}")
        narrative.append("")
    health = snapshot.get("enterprise_health_score", 0)
    posture = "Strong" if health >= 80 else "Stable" if health >= 60 else "Elevated" if health >= 40 else "Critical"
    narrative.append(f"## 1. Executive Posture Summary")
    narrative.append(f"**Enterprise Health Score:** {health}/100 ({posture})")
    narrative.append(f"**Total Risks:** {snapshot.get('critical_count', 0) + snapshot.get('high_count', 0)} critical/high, {snapshot.get('avg_residual', 0):.1f}/25 average residual")
    narrative.append("")
    if comparison:
        narrative.append(f"## 2. Movement Since Last Review")
        if comparison.get("new_risks"):
            narrative.append(f"- **New risks:** {len(comparison['new_risks'])}")
        if comparison.get("closed_risks"):
            narrative.append(f"- **Closed risks:** {len(comparison['closed_risks'])}")
        if comparison.get("worsened_risks"):
            narrative.append(f"- **Worsened risks:** {len(comparison['worsened_risks'])}")
        if comparison.get("improved_risks"):
            narrative.append(f"- **Improved risks:** {len(comparison['improved_risks'])}")
        narrative.append("")
    narrative.append(f"## 3. Concentration Risk Areas")
    narrative.append(f"**Top Division Exposure:** {snapshot.get('top_division', 'N/A')} ({snapshot.get('top_division_pct', 0)}% of enterprise load)")
    narrative.append("")
    narrative.append(f"## 4. Risk Appetite Status (Threshold: {threshold}/25)")
    narrative.append(f"- Within appetite: {snapshot.get('pct_within_appetite', 0)}%")
    narrative.append(f"- Near appetite: {snapshot.get('pct_near_appetite', 0)}%")
    narrative.append(f"- Breached: {snapshot.get('pct_breached', 0)}%")
    narrative.append("")
    narrative.append(f"## 5. Treatment Delivery Confidence")
    narrative.append(f"**Confidence Score:** {snapshot.get('treatment_confidence', 0)}%")
    narrative.append(f"**Ownership coverage:** {snapshot.get('ownership_coverage', 0)}%")
    narrative.append("")
    if snapshot.get("board_risks"):
        narrative.append(f"## 6. Top 5 Board-Attention Risks")
        for risk in snapshot["board_risks"]:
            narrative.append(f"- **{risk['risk_name']}** ({risk['division']}) – Residual: {risk['residual_score']}/25, Owner: {risk['owner']}")
        narrative.append("")
    themes = snapshot.get("emerging_themes", [])
    if themes:
        narrative.append(f"## 7. Emerging Systemic Themes")
        for theme in themes:
            narrative.append(f"- {theme}")
        narrative.append("")
    return "\n".join(narrative)

# =============================================================================
# EXCEL & PDF EXPORTS
# =============================================================================
def style_excel_with_risk_colors(wb):
    level_colors = {
        "Critical": PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid"),
        "High": PatternFill(start_color="FFEDD5", end_color="FFEDD5", fill_type="solid"),
        "Medium": PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid"),
        "Low": PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
    }
    for sheetname in wb.sheetnames:
        ws = wb[sheetname]
        header_row = 1
        level_col_idx = None
        for col_idx, cell in enumerate(ws[header_row], 1):
            if cell.value and "residual_level" in str(cell.value).lower():
                level_col_idx = col_idx
                break
        if level_col_idx:
            for row in range(2, ws.max_row + 1):
                level = ws.cell(row=row, column=level_col_idx).value
                fill = level_colors.get(level, PatternFill())
                for col in range(1, ws.max_column + 1):
                    ws.cell(row=row, column=col).fill = fill

def generate_intelligent_excel_pack(data: Dict[str, Any], narrative: str, correlations: List[Dict], trends: Dict) -> io.BytesIO:
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        summary = pd.DataFrame({
            "Metric": ["Organization", "Report Title", "Period", "Board Date", "Health Score", "Total Risks", "Critical+High", "Avg Residual", "Top Division", "Treatment Confidence", "Within Appetite", "Near Appetite", "Breached"],
            "Value": [
                data.get("company", ""), data.get("report_title", ""), data.get("period", ""), data.get("board_date", ""),
                data.get("enterprise_health_score", 0), data["total_risks"],
                data.get("critical_count", 0) + data.get("high_count", 0),
                f"{data.get('avg_residual', 0):.1f}/25", f"{data.get('top_division', 'N/A')} ({data.get('top_division_pct', 0)}%)",
                f"{data.get('treatment_confidence', 0)}%", f"{data.get('pct_within_appetite', 0)}%",
                f"{data.get('pct_near_appetite', 0)}%", f"{data.get('pct_breached', 0)}%"
            ]
        })
        summary.to_excel(writer, sheet_name="Executive Summary", index=False)
        pd.DataFrame({"Board Narrative": narrative.split("\n")}).to_excel(writer, sheet_name="Board Narrative", index=False)
        data["risks_df"].to_excel(writer, sheet_name="Enterprise Risks", index=False)
        board_risks = data["risks_df"][data["risks_df"]["residual_score"] >= data.get("threshold", 12)].copy()
        if not board_risks.empty:
            board_risks.to_excel(writer, sheet_name="Board Attention", index=False)
        if data.get("division_exposure"):
            div_df = pd.DataFrame(list(data["division_exposure"].items()), columns=["Division", "Exposure"])
            div_df.to_excel(writer, sheet_name="Division Exposure", index=False)
        if data.get("category_exposure"):
            cat_df = pd.DataFrame(list(data["category_exposure"].items()), columns=["Category", "Exposure"])
            cat_df.to_excel(writer, sheet_name="Category Exposure", index=False)
        if st.session_state.parser_audit:
            audit_df = pd.DataFrame([st.session_state.parser_audit])
            audit_df.to_excel(writer, sheet_name="Parser Audit", index=False)
        if correlations:
            corr_df = pd.DataFrame([{
                "Risk A Name": c["risk_a"]["name"],
                "Risk A Division": c["risk_a"]["division"],
                "Risk B Name": c["risk_b"]["name"],
                "Risk B Division": c["risk_b"]["division"],
                "Similarity %": c["similarity"]
            } for c in correlations])
            corr_df.to_excel(writer, sheet_name="Cross-Division Correlations", index=False)
        if trends:
            trends_df = pd.DataFrame([trends])
            trends_df.to_excel(writer, sheet_name="Trend Analysis", index=False)
        briefing_df = pd.DataFrame({"Executive Briefing": [data.get("ai_summary", "")]})
        briefing_df.to_excel(writer, sheet_name="AI Executive Briefing", index=False)

    output.seek(0)
    wb = load_workbook(output)
    style_excel_with_risk_colors(wb)
    styled_output = io.BytesIO()
    wb.save(styled_output)
    styled_output.seek(0)
    return styled_output

def generate_pdf_board_pack(narrative: str, snapshot: Dict, company: str, report_title: str, logo_bytes: bytes = None) -> bytes:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter))
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=18, alignment=TA_CENTER, spaceAfter=20, textColor=colors.HexColor("#0E365C"))
    heading_style = ParagraphStyle('Heading', parent=styles['Heading2'], fontSize=14, spaceBefore=15, spaceAfter=10, textColor=colors.HexColor("#1A5F7A"))
    normal_style = ParagraphStyle('Normal', parent=styles['Normal'], fontSize=10, spaceAfter=6)
    story = []
    if logo_bytes:
        img = Image(io.BytesIO(logo_bytes), width=1.5*inch, height=0.75*inch)
        story.append(img)
    story.append(Paragraph(f"{report_title}", title_style))
    story.append(Paragraph(f"{company}", title_style))
    story.append(Spacer(1, 12))
    health = snapshot.get("enterprise_health_score", 0)
    story.append(Paragraph(f"Enterprise Health Score: {health}/100", heading_style))
    story.append(Spacer(1, 6))
    for line in narrative.split("\n"):
        if line.startswith("#"):
            story.append(Paragraph(line[2:], heading_style))
        elif line.strip():
            story.append(Paragraph(line, normal_style))
        else:
            story.append(Spacer(1, 6))
    if snapshot.get("board_risks"):
        story.append(Spacer(1, 12))
        story.append(Paragraph("Top Board-Attention Risks (Color Coded)", heading_style))
        data = []
        for risk in snapshot["board_risks"][:5]:
            level = "Critical" if risk["residual_score"] >= 20 else "High" if risk["residual_score"] >= 12 else "Medium" if risk["residual_score"] >= 6 else "Low"
            data.append([risk["risk_name"], risk["division"], f"{risk['residual_score']}/25", risk["owner"], level])
        t = Table([["Risk Name", "Division", "Residual", "Owner", "Level"]] + data, colWidths=[200, 100, 60, 100, 60])
        style = [('BACKGROUND', (0,0), (-1,0), colors.HexColor("#0E365C")), ('TEXTCOLOR', (0,0), (-1,0), colors.white)]
        for i, row in enumerate(data, start=1):
            level = row[4]
            if level == "Critical":
                bg = colors.HexColor("#FEE2E2")
            elif level == "High":
                bg = colors.HexColor("#FFEDD5")
            elif level == "Medium":
                bg = colors.HexColor("#FEF3C7")
            else:
                bg = colors.HexColor("#DCFCE7")
            style.append(('BACKGROUND', (0,i), (-1,i), bg))
        t.setStyle(style)
        story.append(t)
    if st.session_state.parser_audit:
        story.append(Spacer(1, 12))
        story.append(Paragraph("Parser Audit Summary", heading_style))
        audit_data = [[k, str(v)] for k, v in st.session_state.parser_audit.items()]
        audit_table = Table(audit_data, colWidths=[150, 150])
        audit_table.setStyle([('BACKGROUND', (0,0), (-1,0), colors.HexColor("#0E365C")), ('TEXTCOLOR', (0,0), (-1,0), colors.white)])
        story.append(audit_table)
    doc.build(story)
    return buffer.getvalue()

# =============================================================================
# DASHBOARD CHARTS
# =============================================================================
def create_category_chart(df: pd.DataFrame) -> go.Figure:
    if df.empty:
        return go.Figure()
    df_plot = df.copy()
    df_plot["category_clean"] = df_plot["category"].apply(lambda x: x.split("/")[0].strip() if isinstance(x, str) else x)
    cat_exposure = df_plot.groupby("category_clean")["residual_score"].sum().sort_values(ascending=False).head(8)
    fig = go.Figure(data=[go.Bar(x=list(cat_exposure.values), y=list(cat_exposure.index), orientation='h', marker_color='#4A90E2')])
    fig.update_layout(title="Risk Exposure by Category", height=350, plot_bgcolor="white", margin=dict(l=10, r=10, t=40, b=10))
    return fig

def create_division_chart(df: pd.DataFrame) -> go.Figure:
    if df.empty:
        return go.Figure()
    df_plot = df.copy()
    df_plot = df_plot[~df_plot["division"].str.match(r"^\d{4}[-/]\d{2}$", na=False)]
    df_plot = df_plot[df_plot["division"] != "Unknown Division"]
    div_exposure = df_plot.groupby("division")["residual_score"].sum().sort_values(ascending=False).head(8)
    fig = go.Figure(data=[go.Bar(x=list(div_exposure.index), y=list(div_exposure.values), marker_color='#F97316')])
    fig.update_layout(title="Risk Exposure by Division", height=350, xaxis_tickangle=-30, plot_bgcolor="white", margin=dict(l=10, r=10, t=40, b=80))
    return fig

def create_appetite_gauge(df: pd.DataFrame, threshold: int, category_appetite: Dict = None) -> go.Figure:
    if df.empty:
        return go.Figure()
    df["appetite"] = df.apply(lambda row: appetite_band(row["residual_score"], threshold, row.get("category", ""), category_appetite), axis=1)
    counts = df["appetite"].value_counts()
    labels = ["Within", "Near", "Breached"]
    values = [counts.get("within appetite", 0), counts.get("near appetite", 0), counts.get("breached", 0) + counts.get("critical breach", 0)]
    fig = go.Figure(data=[go.Pie(labels=labels, values=values, hole=0.4, marker_colors=["#10B981", "#F59E0B", "#EF4444"])])
    fig.update_layout(title="Risk Appetite Status", height=300, margin=dict(l=10, r=10, t=40, b=10))
    return fig

def create_treatment_gauge(confidence: float) -> go.Figure:
    fig = go.Figure(go.Indicator(
        mode="gauge+number+delta",
        value=confidence,
        title={"text": "Treatment Confidence", "font": {"size": 14}},
        domain={"x": [0, 1], "y": [0, 1]},
        gauge={"axis": {"range": [0, 100]}, "bar": {"color": "#6C63FF"},
               "steps": [{"range": [0, 50], "color": "#FEE2E2"}, {"range": [50, 75], "color": "#FEF3C7"}, {"range": [75, 100], "color": "#DCFCE7"}],
               "threshold": {"line": {"color": "red", "width": 4}, "thickness": 0.75, "value": confidence}}))
    fig.update_layout(height=250, margin=dict(l=10, r=10, t=50, b=10))
    return fig

# =============================================================================
# UI COMPONENTS
# =============================================================================
def apply_custom_theme(primary: str, secondary: str) -> None:
    st.markdown(f"""
    <style>
    body, .stApp {{ background-color: #f9fafb; }}
    .stTabs [data-baseweb="tab-list"] {{ gap: 8px; background-color: transparent; }}
    .stTabs [data-baseweb="tab"] {{
        border-radius: 12px 12px 0 0; padding: 12px 24px; background-color: white;
        border: 1px solid #e2e8f0; border-bottom: none; font-weight: 600; color: #475569;
    }}
    .stTabs [aria-selected="true"] {{
        background-color: {primary} !important; color: white !important; border-color: {primary} !important;
    }}
    .stButton > button {{
        background: linear-gradient(145deg, {primary} 0%, {secondary} 100%);
        color: white; border: none; border-radius: 40px; padding: 12px 28px; font-weight: 600;
        transition: all 0.2s; box-shadow: 0 4px 12px rgba(14, 54, 92, 0.2);
    }}
    .stButton > button:hover {{ transform: translateY(-2px); box-shadow: 0 8px 20px rgba(14, 54, 92, 0.3); }}
    .exec-card {{
        background: #ffffff; border-radius: 20px; padding: 22px 24px;
        box-shadow: 0 8px 24px rgba(15, 23, 42, 0.04); border: 1px solid rgba(14, 54, 92, 0.08); margin-bottom: 20px;
    }}
    .exec-card-header {{
        display: flex; align-items: center; margin-bottom: 18px; border-bottom: 1px solid #eef2f6; padding-bottom: 14px;
    }}
    .exec-card-title {{ font-size: 1.2rem; font-weight: 700; color: {primary}; letter-spacing: -0.01em; }}
    .exec-badge {{
        display: inline-block; padding: 6px 14px; border-radius: 40px; font-size: 0.8rem;
        font-weight: 700; letter-spacing: 0.2px; text-transform: uppercase;
    }}
    .exec-metric-row {{ display: flex; justify-content: space-between; margin-bottom: 14px; }}
    .exec-metric-label {{ color: #475569; font-size: 0.95rem; }}
    .exec-metric-value {{ font-weight: 700; color: #0f172a; font-size: 1.1rem; }}
    .exec-divider {{ height: 1px; background: #e2e8f0; margin: 18px 0; }}
    .exec-risk-card {{
        background: #f8fafc; border-radius: 16px; padding: 16px 20px; margin-bottom: 12px; border-left: 6px solid;
    }}
    .exec-hero {{
        background: linear-gradient(145deg, {primary} 0%, {secondary} 100%);
        border-radius: 28px; padding: 28px 32px; color: white; margin-bottom: 24px;
        box-shadow: 0 12px 32px rgba(14, 54, 92, 0.2);
    }}
    .exec-ai-brief {{
        background: #f1f5f9; border-radius: 18px; padding: 20px 24px;
        border-left: 8px solid {primary}; margin-bottom: 24px;
    }}
    .kpi-card {{
        background: white; border-radius: 16px; padding: 18px 16px;
        box-shadow: 0 4px 12px rgba(0,0,0,0.04); border: 1px solid #edf2f7; height: 100%;
    }}
    .kpi-label {{ color: #64748b; font-size: 0.9rem; font-weight: 500; letter-spacing: 0.3px; margin-bottom: 8px; }}
    .kpi-value {{ color: #0f172a; font-size: 2.2rem; font-weight: 700; line-height: 1.2; }}
    .register-table {{
        width: 100%; border-collapse: collapse; font-family: -apple-system, BlinkMacSystemFont, sans-serif; font-size: 13px;
    }}
    .register-table th {{
        background-color: #f3f4f6; color: #1f2937; font-weight: 600; border-bottom: 2px solid #d1d5db;
        padding: 10px 8px; text-align: left; border-right: 1px solid #e5e7eb;
    }}
    .register-table td {{
        padding: 8px; border-right: 1px solid #e5e7eb; border-bottom: 1px solid #e5e7eb; vertical-align: top;
    }}
    </style>
    """, unsafe_allow_html=True)

def render_parser_audit_panel():
    if st.session_state.parser_audit:
        with st.expander("🔍 Extraction Quality Dashboard", expanded=False):
            audit = st.session_state.parser_audit
            col1, col2, col3, col4 = st.columns(4)
            col1.metric("Files Processed", audit.get("total_files", 0))
            col2.metric("Rows Scanned", audit.get("rows_scanned", 0))
            col3.metric("Rows Accepted", audit.get("rows_accepted", 0))
            col4.metric("Acceptance Rate", f"{audit.get('acceptance_rate', 0)}%")
            col5, col6, col7 = st.columns(3)
            col5.metric("Total Risks", audit.get("total_risks", 0))
            col6.metric("Duplicates Removed", audit.get("duplicates_removed", 0))
            col7.metric("Clusters Formed", audit.get("clusters_formed", 0))

def render_sidebar():
    with st.sidebar:
        st.markdown("## 🛡️ RiskForge")
        st.caption(f"**Tier:** {st.session_state.tier.upper()}")
        uploaded_logo = st.file_uploader("Company Logo", type=["png", "jpg", "jpeg"], key="logo_upload")
        if uploaded_logo:
            st.session_state.logo_bytes = uploaded_logo.getvalue()
        if st.session_state.tier == "free":
            st.markdown("---")
            st.markdown("### 🚀 Upgrade")
            st.markdown("""
**Professional** – $29/month or $99/year  
- Full board pack  
- AI briefing  
- Heatmaps  
- Category/Division charts  
""")
            col1, col2 = st.columns(2)
            with col1:
                if st.button("$29/mo", key="pro_monthly"):
                    if not stripe or not STRIPE_SECRET_KEY:
                        st.error("Stripe is not configured.")
                    elif not STRIPE_PRICE_ID_PRO_MONTHLY:
                        st.error("Professional monthly price ID missing.")
                    else:
                        try:
                            session = stripe.checkout.Session.create(
                                payment_method_types=["card"],
                                line_items=[{"price": STRIPE_PRICE_ID_PRO_MONTHLY, "quantity": 1}],
                                mode="subscription",
                                success_url=APP_URL + "?success_pro_monthly=true",
                                cancel_url=APP_URL,
                            )
                            st.markdown(f"<a href='{session.url}' target='_blank'>Pay</a>", unsafe_allow_html=True)
                        except Exception as e:
                            st.error(f"Stripe error: {e}")
            with col2:
                if st.button("$99/yr", key="pro_annual"):
                    if not stripe or not STRIPE_SECRET_KEY:
                        st.error("Stripe is not configured.")
                    elif not STRIPE_PRICE_ID_PRO_ANNUAL:
                        st.error("Professional annual price ID missing.")
                    else:
                        try:
                            session = stripe.checkout.Session.create(
                                payment_method_types=["card"],
                                line_items=[{"price": STRIPE_PRICE_ID_PRO_ANNUAL, "quantity": 1}],
                                mode="subscription",
                                success_url=APP_URL + "?success_pro_annual=true",
                                cancel_url=APP_URL,
                            )
                            st.markdown(f"<a href='{session.url}' target='_blank'>Pay</a>", unsafe_allow_html=True)
                        except Exception as e:
                            st.error(f"Stripe error: {e}")
            st.markdown("")
            st.markdown("""
**Enterprise** – $99/month or $299/year  
- Branded PDF board pack  
- Committee-ready exports  
- White-label reports  
- Priority support  
- Custom appetite thresholds  
""")
            col1, col2 = st.columns(2)
            with col1:
                if st.button("$99/mo", key="ent_monthly"):
                    if not stripe or not STRIPE_SECRET_KEY:
                        st.error("Stripe is not configured.")
                    elif not STRIPE_PRICE_ID_ENT_MONTHLY:
                        st.error("Enterprise monthly price ID missing.")
                    else:
                        try:
                            session = stripe.checkout.Session.create(
                                payment_method_types=["card"],
                                line_items=[{"price": STRIPE_PRICE_ID_ENT_MONTHLY, "quantity": 1}],
                                mode="subscription",
                                success_url=APP_URL + "?success_ent_monthly=true",
                                cancel_url=APP_URL,
                            )
                            st.markdown(f"<a href='{session.url}' target='_blank'>Pay</a>", unsafe_allow_html=True)
                        except Exception as e:
                            st.error(f"Stripe error: {e}")
            with col2:
                if st.button("$299/yr", key="ent_annual"):
                    if not stripe or not STRIPE_SECRET_KEY:
                        st.error("Stripe is not configured.")
                    elif not STRIPE_PRICE_ID_ENT_ANNUAL:
                        st.error("Enterprise annual price ID missing.")
                    else:
                        try:
                            session = stripe.checkout.Session.create(
                                payment_method_types=["card"],
                                line_items=[{"price": STRIPE_PRICE_ID_ENT_ANNUAL, "quantity": 1}],
                                mode="subscription",
                                success_url=APP_URL + "?success_ent_annual=true",
                                cancel_url=APP_URL,
                            )
                            st.markdown(f"<a href='{session.url}' target='_blank'>Pay</a>", unsafe_allow_html=True)
                        except Exception as e:
                            st.error(f"Stripe error: {e}")
            st.markdown("---")
            code = st.text_input("Unlock code", type="password", placeholder="Enter code")
            col_apply, _ = st.columns([1, 2])
            with col_apply:
                apply_btn = st.button("Apply Code", use_container_width=True)

            if apply_btn:
                if code == PRO_UNLOCK_CODE:
                    st.session_state.tier = "professional"
                    st.success("✅ Professional unlocked!")
                    st.rerun()
                elif code == ENT_UNLOCK_CODE:
                    st.session_state.tier = "enterprise"
                    st.success("✅ Enterprise unlocked!")
                    st.rerun()
                elif code:
                    st.error("❌ Invalid unlock code")
        else:
            st.success(f"✅ {st.session_state.tier.upper()} active")
        st.markdown("---")
        st.session_state.org_name = st.text_input("Organization Name", st.session_state.org_name)
        st.session_state.report_title = st.text_input("Report Title", st.session_state.report_title)
        st.session_state.board_threshold = st.slider("Global Board Threshold", 0, 25, st.session_state.board_threshold)
        st.session_state.default_residual_score = st.slider("Default residual score (when missing)", 0, 25, st.session_state.default_residual_score)
        if st.session_state.tier == "enterprise":
            st.markdown("**Per-Category Appetite (Enterprise)**")
            categories = ["Strategic", "Financial", "Operational", "ICT/Cyber", "Compliance/Legal", "People/HR", "Health/Safety", "Reputational", "Environmental"]
            for cat in categories:
                current = st.session_state.category_appetite.get(cat, st.session_state.board_threshold)
                new_val = st.number_input(f"{cat}", min_value=0, max_value=25, value=current, key=f"appetite_{cat}")
                if new_val != current:
                    st.session_state.category_appetite[cat] = new_val
        st.session_state.primary_color = st.color_picker("Primary Color", st.session_state.primary_color)
        st.session_state.secondary_color = st.color_picker("Secondary Color", st.session_state.secondary_color)
        st.checkbox("🔧 Debug Mode", key="debug_mode")
        if GEMINI_AVAILABLE:
            st.checkbox("🤖 Force Gemini Extraction", key="force_gemini", value=True)

# =============================================================================
# MAIN APP
# =============================================================================
def main():
    apply_custom_theme(st.session_state.primary_color, st.session_state.secondary_color)
    render_sidebar()

    # Hero state is now correctly synchronized below after processing

    if GEMINI_AVAILABLE:
        st.success("✅ Gemini API is configured and ready")
    else:
        st.error("❌ Gemini API key not found")

    if st.session_state.tier == "free":
        st.info("🔓 Free tier: 1 file, preview only. Upgrade for full features.")

    max_files = 1 if st.session_state.tier == "free" else 999
    uploaded_files = st.file_uploader("Upload risk registers (Excel/CSV)", accept_multiple_files=True, type=["xlsx", "xls", "csv"])

    if len(uploaded_files) > max_files:
        st.error(f"Free tier allows {max_files} file(s).")
        uploaded_files = uploaded_files[:max_files]

    if st.button("🚀 Generate Board Pack", type="primary", use_container_width=True):
        if not uploaded_files:
            st.warning("Please upload files.")
        else:
            with st.spinner("Processing risk registers..."):
                df_all, debug_list = parse_all_files(uploaded_files, st.session_state.tier, st.session_state.default_residual_score)
                if df_all.empty:
                    st.error("No valid risk data found.")
                    if st.session_state.debug_mode:
                        st.subheader("🔧 Parser Debug Information")
                        for i, debug in enumerate(debug_list):
                            st.markdown(f"**File {i+1}**")
                            st.json(debug)
                else:
                    st.success(f"✅ {len(df_all)} risks processed")
                    if st.session_state.debug_mode:
                        with st.expander("🔧 Parsed Risks Preview"):
                            st.dataframe(df_all.head(20))

                    category_appetite = st.session_state.category_appetite if st.session_state.tier == "enterprise" else None
                    snapshot = build_intelligence_snapshot(df_all, st.session_state.board_threshold, category_appetite)
                    comparison = {}
                    trends = {}
                    if st.session_state.history:
                        comparison = compare_snapshots(snapshot, st.session_state.history[-1])
                        trends = analyze_trends(snapshot, st.session_state.history[-1])

                    correlations = find_cross_division_correlations(df_all)
                    recommendations = recommend_appetite_thresholds(df_all, st.session_state.board_threshold)

                    ai_summary = ""
                    if st.session_state.tier != "free" and GEMINI_AVAILABLE:
                        ai_summary = ai_executive_briefing(snapshot, correlations, recommendations, st.session_state.org_name)

                    narrative = generate_board_narrative(snapshot, comparison, st.session_state.board_threshold, st.session_state.org_name, st.session_state.report_title, ai_summary)
                    st.session_state.rf_data = {
                        "risks_df": df_all,
                        "total_risks": len(df_all),
                        "company": st.session_state.org_name,
                        "report_title": st.session_state.report_title,
                        "period": f"Q{((datetime.now().month-1)//3)+1} {datetime.now().year}",
                        "board_date": datetime.now().strftime("%B %d, %Y"),
                        "threshold": st.session_state.board_threshold,
                        "critical_count": snapshot.get("critical_count", 0),
                        "high_count": snapshot.get("high_count", 0),
                        "avg_residual": snapshot.get("avg_residual", 0),
                        "avg_inherent": snapshot.get("avg_inherent", 0),
                        "enterprise_health_score": snapshot.get("enterprise_health_score", 0),
                        "treatment_confidence": snapshot.get("treatment_confidence", 0),
                        "top_division": snapshot.get("top_division", "N/A"),
                        "top_division_pct": snapshot.get("top_division_pct", 0),
                        "division_exposure": snapshot.get("division_exposure", {}),
                        "category_exposure": snapshot.get("category_exposure", {}),
                        "ownership_coverage": snapshot.get("ownership_coverage", 0),
                        "pct_within_appetite": snapshot.get("pct_within_appetite", 0),
                        "pct_near_appetite": snapshot.get("pct_near_appetite", 0),
                        "pct_breached": snapshot.get("pct_breached", 0),
                        "emerging_themes": snapshot.get("emerging_themes", []),
                        "board_risks": snapshot.get("board_risks", []),
                        "narrative": narrative,
                        "comparison": comparison,
                        "trends": trends,
                        "correlations": correlations,
                        "recommendations": recommendations,
                        "ai_summary": ai_summary
                    }
                    st.session_state.history.append(snapshot)
                    if len(st.session_state.history) > 4:
                        st.session_state.history = st.session_state.history[-4:]
                    if st.session_state.tier != "free":
                        excel_data = generate_intelligent_excel_pack(st.session_state.rf_data, narrative, correlations, trends)
                        st.download_button("📥 Excel Board Pack", excel_data, file_name=f"RiskForge_{datetime.now().strftime('%Y%m%d')}.xlsx")
                        if st.session_state.tier == "enterprise":
                            pdf_data = generate_pdf_board_pack(narrative, snapshot, st.session_state.org_name, st.session_state.report_title, st.session_state.logo_bytes)
                            st.download_button("📥 PDF Board Pack (Enterprise)", pdf_data, file_name=f"BoardPack_{datetime.now().strftime('%Y%m%d')}.pdf")
                    else:
                        st.info("📌 Upgrade to Professional/Enterprise to download board packs.")
                    st.rerun()

    # Hero header (now correctly after processing, uses st.session_state.rf_data)
    if st.session_state.rf_data:
        health = st.session_state.rf_data.get("enterprise_health_score", 0)
    else:
        health = 0

    if health >= 80:
        posture, posture_color, posture_bg = "Strong", "#10b981", "#d1fae5"
    elif health >= 60:
        posture, posture_color, posture_bg = "Stable", "#3b82f6", "#dbeafe"
    elif health >= 40:
        posture, posture_color, posture_bg = "Elevated", "#f59e0b", "#fef3c7"
    else:
        posture, posture_color, posture_bg = "Critical", "#ef4444", "#fee2e2"

    col_logo, col_hero = st.columns([0.5, 5])
    with col_logo:
        if st.session_state.logo_bytes:
            st.image(st.session_state.logo_bytes, width=60)
        else:
            st.markdown("### 🛡️")
    with col_hero:
        st.markdown(f"""
        <div class="exec-hero" style="margin-top: 0;">
            <div style="display: flex; justify-content: space-between; align-items: flex-start;">
                <div>
                    <div style="font-size: 14px; opacity: 0.85; letter-spacing: 0.5px;">
                        {st.session_state.org_name.upper()}
                    </div>
                    <div style="font-size: 36px; font-weight: 700; margin-top: 8px; line-height: 1.1;">
                        {st.session_state.report_title}
                    </div>
                    <div style="margin-top: 12px; font-size: 16px; opacity: 0.9;">
                        Reporting Period: Q{((datetime.now().month-1)//3)+1} {datetime.now().year} • Board Date: {datetime.now().strftime('%B %d, %Y')}
                    </div>
                </div>
                <div style="text-align: right;">
                    <div style="font-size: 14px; opacity: 0.8;">ENTERPRISE HEALTH</div>
                    <div style="font-size: 56px; font-weight: 800; line-height: 1;">{health}</div>
                    <div style="background: {posture_bg}; color: {posture_color}; padding: 6px 16px; border-radius: 40px; font-weight: 700; font-size: 16px; margin-top: 8px; display: inline-block;">
                        {posture} POSTURE
                    </div>
                </div>
            </div>
        </div>
        """, unsafe_allow_html=True)

    render_parser_audit_panel()

    if st.session_state.rf_data:
        data = st.session_state.rf_data
        df = data["risks_df"]
        tab1, tab2, tab3, tab4, tab5 = st.tabs(["📊 Dashboard", "📑 Register", "🧠 Intelligence", "📈 Trends", "📤 Export"])

        with tab1:
            st.subheader("Executive Dashboard")
            col1, col2, col3, col4, col5 = st.columns(5)
            with col1:
                st.markdown(f"""
                <div class="kpi-card">
                    <div class="kpi-label">📋 Total Risks</div>
                    <div class="kpi-value">{data['total_risks']}</div>
                </div>
                """, unsafe_allow_html=True)
            with col2:
                st.markdown(f"""
                <div class="kpi-card">
                    <div class="kpi-label">⚠️ Critical + High</div>
                    <div class="kpi-value">{data['critical_count'] + data['high_count']}</div>
                </div>
                """, unsafe_allow_html=True)
            with col3:
                health_val = data['enterprise_health_score']
                health_color = "#10b981" if health_val >= 80 else "#3b82f6" if health_val >= 60 else "#f59e0b" if health_val >= 40 else "#ef4444"
                st.markdown(f"""
                <div class="kpi-card">
                    <div class="kpi-label">❤️ Health Score</div>
                    <div class="kpi-value" style="color: {health_color};">{health_val}/100</div>
                </div>
                """, unsafe_allow_html=True)
            with col4:
                conf = data.get('treatment_confidence', 0)
                conf_color = "#10b981" if conf >= 75 else "#f59e0b" if conf >= 50 else "#ef4444"
                st.markdown(f"""
                <div class="kpi-card">
                    <div class="kpi-label">🛠️ Treatment Confidence</div>
                    <div class="kpi-value" style="color: {conf_color};">{conf}%</div>
                </div>
                """, unsafe_allow_html=True)
            with col5:
                st.markdown(f"""
                <div class="kpi-card">
                    <div class="kpi-label">🏢 Top Division</div>
                    <div class="kpi-value" style="font-size: 1.3rem;">{data['top_division']}</div>
                    <div style="color: #64748b; font-size: 0.8rem;">{data['top_division_pct']}% of load</div>
                </div>
                """, unsafe_allow_html=True)

            col_chart1, col_chart2 = st.columns(2)
            with col_chart1:
                if not df.empty:
                    fig_cat = create_category_chart(df)
                    st.plotly_chart(fig_cat, use_container_width=True)
            with col_chart2:
                if not df.empty:
                    fig_div = create_division_chart(df)
                    st.plotly_chart(fig_div, use_container_width=True)
            col_app, col_treat = st.columns(2)
            with col_app:
                if not df.empty:
                    fig_app = create_appetite_gauge(df, data["threshold"], st.session_state.category_appetite if st.session_state.tier == "enterprise" else None)
                    st.plotly_chart(fig_app, use_container_width=True)
            with col_treat:
                fig_treat = create_treatment_gauge(data.get("treatment_confidence", 0))
                st.plotly_chart(fig_treat, use_container_width=True)
            st.subheader("Risk Appetite Status")
            st.progress(data.get("pct_within_appetite", 0) / 100)
            st.caption(f"Within: {data.get('pct_within_appetite', 0)}% | Near: {data.get('pct_near_appetite', 0)}% | Breached: {data.get('pct_breached', 0)}%")

        with tab2:
            st.subheader("Enterprise Risk Register")

            register_cols = [
                "division", "risk_no", "risk_name", "risk_statement", "cause", "category",
                "impact_score", "likelihood_score", "inherent_score", "controls",
                "control_effectiveness", "owner", "strategy", "treatment_plan",
                "due_date_raw", "residual_score", "residual_level",
            ]

            display_df = df.copy()
            display_df = display_df[[c for c in register_cols if c in display_df.columns]]

            display_df = display_df.rename(columns={
                "division": "Division/Dept",
                "risk_no": "Risk No",
                "risk_name": "Risk Description",
                "risk_statement": "Risk Definition/Statement",
                "cause": "Cause",
                "category": "Risk Category",
                "impact_score": "Impact (1-5)",
                "likelihood_score": "Likelihood (1-5)",
                "inherent_score": "Inherent Risk",
                "controls": "Controls",
                "control_effectiveness": "Control Effectiveness",
                "owner": "Risk Owner",
                "strategy": "Risk Strategy",
                "treatment_plan": "Risk Treatment/Action Plan",
                "due_date_raw": "Treatment Due Date",
                "residual_score": "Residual Risk",
                "residual_level": "Residual Level",
            })

            threshold = data.get("threshold", st.session_state.board_threshold)
            category_appetite = st.session_state.category_appetite if st.session_state.tier == "enterprise" else {}

            html_table = '<table class="register-table">'
            html_table += '<thead><tr>'
            for col in display_df.columns:
                html_table += f'<th>{col}</th>'
            html_table += '</tr></thead><tbody>'

            for _, row in display_df.iterrows():
                band = appetite_band(row["Residual Risk"], threshold, row.get("Risk Category", ""), category_appetite)
                if band in ["breached", "critical breach"]:
                    row_bg = "#fee2e2"
                elif band == "near appetite":
                    row_bg = "#fef3c7"
                elif band == "within appetite":
                    row_bg = "#dcfce7"
                else:
                    row_bg = "#ffffff"

                html_table += f'<tr style="background-color: {row_bg};">'
                for col_name, value in row.items():
                    cell_style = ""
                    if col_name in ["Inherent Risk", "Residual Risk"]:
                        try:
                            score = float(value)
                            if score >= 20:
                                cell_style = "background-color: #dc2626; color: white; font-weight: 700;"
                            elif score >= 12:
                                cell_style = "background-color: #f59e0b; color: white; font-weight: 700;"
                            elif score >= 6:
                                cell_style = "background-color: #fef3c7; color: #92400e; font-weight: 600;"
                            else:
                                cell_style = "background-color: #dcfce7; color: #166534; font-weight: 600;"
                        except:
                            pass
                    elif col_name in ["Impact (1-5)", "Likelihood (1-5)"]:
                        try:
                            val = float(value)
                            if val >= 4:
                                cell_style = "background-color: #fca5a5; font-weight: 600;"
                            elif val >= 3:
                                cell_style = "background-color: #fde68a; font-weight: 600;"
                        except:
                            pass

                    display_val = str(value) if pd.notna(value) else ""
                    html_table += f'<td style="{cell_style}">{display_val}</td>'
                html_table += '</tr>'
            html_table += '</tbody></table>'

            st.markdown(html_table, unsafe_allow_html=True)

            with st.expander("🔍 Risk Lineage & Confidence"):
                lineage_df = df[["risk_name", "source_file", "source_sheet", "source_row", "acceptance_score", "acceptance_reason", "parser_confidence"]].copy()
                st.dataframe(lineage_df, use_container_width=True)

        with tab3:
            if data.get("ai_summary"):
                st.markdown(f"""
                <div class="exec-ai-brief">
                    <div style="display: flex; align-items: center; margin-bottom: 12px;">
                        <span style="font-size: 20px; margin-right: 10px;">🤖</span>
                        <span style="font-weight: 700; font-size: 18px; color: #0E365C;">AI Executive Briefing</span>
                    </div>
                    <div style="font-size: 16px; color: #1e293b; line-height: 1.6;">
                        {data['ai_summary']}
                    </div>
                </div>
                """, unsafe_allow_html=True)

            col_left, col_right = st.columns([1.4, 1])

            with col_left:
                st.markdown(f"""
                <div class="exec-card">
                    <div class="exec-card-header">
                        <span class="exec-card-title">📋 Executive Posture</span>
                    </div>
                    <div class="exec-metric-row">
                        <span class="exec-metric-label">Total Risks in Register</span>
                        <span class="exec-metric-value">{data['total_risks']}</span>
                    </div>
                    <div class="exec-metric-row">
                        <span class="exec-metric-label">Critical + High Risks</span>
                        <span class="exec-metric-value">{data['critical_count'] + data['high_count']}</span>
                    </div>
                    <div class="exec-metric-row">
                        <span class="exec-metric-label">Average Residual Score</span>
                        <span class="exec-metric-value">{data.get('avg_residual', 0):.1f} / 25</span>
                    </div>
                    <div class="exec-metric-row">
                        <span class="exec-metric-label">Average Inherent Score</span>
                        <span class="exec-metric-value">{data.get('avg_inherent', 0):.1f} / 25</span>
                    </div>
                    <div class="exec-divider"></div>
                    <div class="exec-metric-row">
                        <span class="exec-metric-label">Ownership Coverage</span>
                        <span class="exec-metric-value">{data.get('ownership_coverage', 0)}%</span>
                    </div>
                </div>
                """, unsafe_allow_html=True)

                comp = data.get("comparison", {})
                if comp:
                    movement_html = ""
                    if comp.get("new_risks"):
                        movement_html += f"""
                        <div class="exec-metric-row">
                            <span class="exec-metric-label">🆕 New Risks</span>
                            <span class="exec-metric-value">{len(comp['new_risks'])}</span>
                        </div>
                        """
                    if comp.get("closed_risks"):
                        movement_html += f"""
                        <div class="exec-metric-row">
                            <span class="exec-metric-label">✅ Closed Risks</span>
                            <span class="exec-metric-value">{len(comp['closed_risks'])}</span>
                        </div>
                        """
                    if comp.get("worsened_risks"):
                        movement_html += f"""
                        <div class="exec-metric-row">
                            <span class="exec-metric-label">📈 Worsened Risks</span>
                            <span class="exec-metric-value">{len(comp['worsened_risks'])}</span>
                        </div>
                        """
                    if comp.get("improved_risks"):
                        movement_html += f"""
                        <div class="exec-metric-row">
                            <span class="exec-metric-label">📉 Improved Risks</span>
                            <span class="exec-metric-value">{len(comp['improved_risks'])}</span>
                        </div>
                        """
                    health_delta = comp.get("health_delta", 0)
                    if health_delta != 0:
                        delta_text = f"+{health_delta:.1f}" if health_delta > 0 else f"{health_delta:.1f}"
                        movement_html += f"""
                        <div class="exec-metric-row">
                            <span class="exec-metric-label">Health Score Change</span>
                            <span class="exec-metric-value">{delta_text} pts</span>
                        </div>
                        """
                    appetite_delta = comp.get("appetite_delta", 0)
                    if appetite_delta != 0:
                        delta_text = f"+{appetite_delta:.1f}%" if appetite_delta > 0 else f"{appetite_delta:.1f}%"
                        movement_html += f"""
                        <div class="exec-metric-row">
                            <span class="exec-metric-label">Appetite Breach Change</span>
                            <span class="exec-metric-value">{delta_text}</span>
                        </div>
                        """

                    if movement_html:
                        st.markdown(f"""
                        <div class="exec-card">
                            <div class="exec-card-header">
                                <span class="exec-card-title">📊 Movement Since Last Review</span>
                            </div>
                            {movement_html}
                        </div>
                        """, unsafe_allow_html=True)

                st.markdown(f"""
                <div class="exec-card">
                    <div class="exec-card-header">
                        <span class="exec-card-title">🎯 Concentration Risk Areas</span>
                    </div>
                    <div class="exec-metric-row">
                        <span class="exec-metric-label">Top Division Exposure</span>
                        <span class="exec-metric-value">{data.get('top_division', 'N/A')}</span>
                    </div>
                    <div class="exec-metric-row">
                        <span class="exec-metric-label">Percentage of Load</span>
                        <span class="exec-metric-value">{data.get('top_division_pct', 0)}%</span>
                    </div>
                </div>
                """, unsafe_allow_html=True)

            with col_right:
                within = data.get('pct_within_appetite', 0)
                near = data.get('pct_near_appetite', 0)
                breached = data.get('pct_breached', 0)

                st.markdown(f"""
                <div class="exec-card">
                    <div class="exec-card-header">
                        <span class="exec-card-title">⚖️ Risk Appetite Status</span>
                    </div>
                    <div style="margin-bottom: 16px;">
                        <div style="display: flex; align-items: center; margin-bottom: 10px;">
                            <div style="width: 12px; height: 12px; border-radius: 4px; background: #10b981; margin-right: 10px;"></div>
                            <span class="exec-metric-label" style="flex:1;">Within Appetite</span>
                            <span class="exec-metric-value">{within}%</span>
                        </div>
                        <div style="display: flex; align-items: center; margin-bottom: 10px;">
                            <div style="width: 12px; height: 12px; border-radius: 4px; background: #f59e0b; margin-right: 10px;"></div>
                            <span class="exec-metric-label" style="flex:1;">Near Appetite</span>
                            <span class="exec-metric-value">{near}%</span>
                        </div>
                        <div style="display: flex; align-items: center;">
                            <div style="width: 12px; height: 12px; border-radius: 4px; background: #ef4444; margin-right: 10px;"></div>
                            <span class="exec-metric-label" style="flex:1;">Breached</span>
                            <span class="exec-metric-value">{breached}%</span>
                        </div>
                    </div>
                    <div class="exec-divider"></div>
                    <div class="exec-metric-row">
                        <span class="exec-metric-label">Global Threshold</span>
                        <span class="exec-metric-value">{data.get('threshold', st.session_state.board_threshold)}/25</span>
                    </div>
                </div>
                """, unsafe_allow_html=True)

                confidence = data.get('treatment_confidence', 0)
                conf_color = "#10b981" if confidence >= 75 else "#f59e0b" if confidence >= 50 else "#ef4444"
                st.markdown(f"""
                <div class="exec-card">
                    <div class="exec-card-header">
                        <span class="exec-card-title">🛠️ Treatment Delivery Confidence</span>
                    </div>
                    <div style="text-align: center; margin: 10px 0;">
                        <span style="font-size: 48px; font-weight: 800; color: {conf_color};">{confidence}%</span>
                    </div>
                    <div class="exec-metric-row">
                        <span class="exec-metric-label">Ownership Coverage</span>
                        <span class="exec-metric-value">{data.get('ownership_coverage', 0)}%</span>
                    </div>
                </div>
                """, unsafe_allow_html=True)

                themes = data.get('emerging_themes', [])
                if themes:
                    themes_html = "".join([f'<span class="exec-badge" style="background:#e0e7ff; color:#3730a3; margin-right:8px; margin-bottom:8px;">{theme}</span>' for theme in themes])
                    st.markdown(f"""
                    <div class="exec-card">
                        <div class="exec-card-header">
                            <span class="exec-card-title">🔍 Emerging Systemic Themes</span>
                        </div>
                        <div style="margin-top: 8px;">
                            {themes_html}
                        </div>
                    </div>
                    """, unsafe_allow_html=True)

            correlations = data.get("correlations", [])
            if correlations:
                with st.expander("🔗 Cross‑Division Risk Correlations", expanded=True):
                    for corr in correlations[:5]:
                        st.markdown(f"""
                        <div style="background:#f1f5f9; border-radius:12px; padding:12px; margin-bottom:8px;">
                            <div style="display:flex; align-items:center; gap:12px;">
                                <div style="flex:1;"><b>{corr['risk_a']['name']}</b><br><small>{corr['risk_a']['division']}</small></div>
                                <div style="background:#0E365C; color:white; padding:4px 12px; border-radius:20px;">{corr['similarity']}% match</div>
                                <div style="flex:1;"><b>{corr['risk_b']['name']}</b><br><small>{corr['risk_b']['division']}</small></div>
                            </div>
                        </div>
                        """, unsafe_allow_html=True)

            recommendations = data.get("recommendations", {})
            if recommendations:
                with st.expander("🎯 Appetite Threshold Recommendations", expanded=False):
                    rec = recommendations.get("global", {})
                    st.markdown(f"**Recommended Global Threshold:** {rec.get('recommended', 'N/A')} (current: {data['threshold']})")
                    st.caption(rec.get('reason', ''))

            st.markdown("""
            <div style="margin-top: 24px;">
                <div class="exec-card-header" style="margin-bottom: 16px;">
                    <span class="exec-card-title" style="font-size: 1.4rem;">⚠️ Top 5 Board‑Attention Risks</span>
                </div>
            </div>
            """, unsafe_allow_html=True)

            board_risks = data.get('board_risks', [])
            for risk in board_risks[:5]:
                score = risk['residual_score']
                if score >= 20:
                    border_color = "#dc2626"
                    score_bg = "#fee2e2"
                    score_color = "#991b1b"
                    level = "Critical"
                elif score >= 12:
                    border_color = "#ea580c"
                    score_bg = "#ffedd5"
                    score_color = "#9a3412"
                    level = "High"
                elif score >= 6:
                    border_color = "#ca8a04"
                    score_bg = "#fef9c3"
                    score_color = "#854d0e"
                    level = "Medium"
                else:
                    border_color = "#16a34a"
                    score_bg = "#dcfce7"
                    score_color = "#166534"
                    level = "Low"

                st.markdown(f"""
                <div class="exec-risk-card" style="border-left-color: {border_color};">
                    <div style="display: flex; justify-content: space-between; align-items: flex-start;">
                        <div style="flex: 1;">
                            <div style="font-size: 18px; font-weight: 700; color: #0E365C; margin-bottom: 6px;">
                                {risk['risk_name']}
                            </div>
                            <div style="font-size: 14px; color: #475569; margin-bottom: 8px;">
                                {risk['division']} • Owner: {risk['owner']} • Category: {risk.get('category', 'Uncategorised')}
                            </div>
                        </div>
                        <div style="margin-left: 20px; text-align: center;">
                            <div style="background: {score_bg}; color: {score_color}; padding: 8px 16px; border-radius: 20px; font-weight: 800; font-size: 20px;">
                                {score}/25
                            </div>
                            <div style="font-size: 12px; font-weight: 600; margin-top: 4px; color: {border_color};">{level}</div>
                        </div>
                    </div>
                </div>
                """, unsafe_allow_html=True)

            if not board_risks:
                st.info("No board‑level risks identified in this submission.")

            with st.expander("📄 View Full Board Narrative (Text Version)"):
                st.markdown(data.get("narrative", "No narrative available"))

        with tab4:
            st.subheader("Quarter Comparison & Trends")
            if data.get("comparison"):
                comp = data["comparison"]
                if comp.get("new_risks"):
                    st.markdown(f"**🆕 New Risks:** {len(comp['new_risks'])}")
                if comp.get("closed_risks"):
                    st.markdown(f"**✅ Closed Risks:** {len(comp['closed_risks'])}")
                if comp.get("worsened_risks"):
                    st.markdown(f"**📈 Worsened Risks:** {len(comp['worsened_risks'])}")
                    for r in comp["worsened_risks"][:5]:
                        st.markdown(f"- {r['name']} (+{r['delta']})")
                if comp.get("improved_risks"):
                    st.markdown(f"**📉 Improved Risks:** {len(comp['improved_risks'])}")
                if comp.get("health_delta", 0) != 0:
                    direction = "improved" if comp["health_delta"] > 0 else "declined"
                    st.markdown(f"**Health Score Change:** {abs(comp['health_delta']):.1f} points {direction}")
                if comp.get("appetite_delta", 0) != 0:
                    direction = "worsened" if comp["appetite_delta"] > 0 else "improved"
                    st.markdown(f"**Appetite Breach Change:** {abs(comp['appetite_delta']):.1f}% {direction}")
            else:
                st.info("Upload another register next quarter to see trend analysis.")

            trends = data.get("trends", {})
            if trends:
                with st.expander("📈 Advanced Trend Analysis", expanded=False):
                    if trends.get("health_trend"):
                        st.markdown(f"**Health Trend:** {trends['health_trend']} ({trends.get('health_delta', 0):+.1f} pts)")
                    if trends.get("breach_trend"):
                        st.markdown(f"**Appetite Breach Trend:** {trends['breach_trend']}")
                    if trends.get("fastest_growing_category"):
                        fgc = trends["fastest_growing_category"]
                        st.markdown(f"**Fastest Growing Risk Category:** {fgc['category']} (+{fgc['growth_pct']}%)")

        with tab5:
            st.subheader("Export Options")
            if st.session_state.tier != "free":
                st.success("✅ Professional/Enterprise tier – full export available above.")
                if st.session_state.tier == "enterprise":
                    st.caption("Enterprise tier includes branded PDF board packs with your organization's logo and custom report titles.")
            else:
                st.info("📌 Upgrade to Professional to download board packs.")

if __name__ == "__main__":
    main()